"""
检查结果导出器
支持导出为Excel、HTML格式
"""
import json
from datetime import datetime
from pathlib import Path
from typing import List
from docx import Document
from docx.shared import Inches, Pt, RGBColor
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models.check_result import DocumentCheckResult, CheckResult


class ResultExporter:
    """结果导出器"""
    
    @staticmethod
    def export_to_excel(doc_result: DocumentCheckResult, output_path: str):
        """
        导出为Excel格式
        
        Args:
            doc_result: 文档检查结果
            output_path: 输出文件路径
        """
        wb = openpyxl.Workbook()
        
        # 创建概览工作表
        ws_summary = wb.active
        ws_summary.title = "检查概览"
        
        # 设置样式
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 添加标题
        ws_summary['A1'] = "Word文档检查报告"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')
        
        # 添加文档信息
        ws_summary['A3'] = "文档路径："
        ws_summary['B3'] = doc_result.document_path
        ws_summary.merge_cells('B3:D3')
        
        ws_summary['A4'] = "检查时间："
        ws_summary['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary.merge_cells('B4:D4')
        
        # 添加统计信息
        ws_summary['A6'] = "检查统计"
        ws_summary['A6'].font = Font(bold=True, size=14)
        
        stats_data = [
            ["总检查项", doc_result.total_rules],
            ["通过项", doc_result.passed_rules],
            ["失败项", doc_result.failed_rules],
            ["警告项", doc_result.warning_rules],
            ["通过率", f"{doc_result.get_pass_rate():.1f}%"],
            ["发现问题", doc_result.get_total_issues()]
        ]
        
        for i, (label, value) in enumerate(stats_data, start=7):
            ws_summary[f'A{i}'] = label
            ws_summary[f'B{i}'] = value
        
        # 创建详细结果工作表
        ws_details = wb.create_sheet("详细结果")
        
        # 表头
        headers = ["序号", "检查类别", "检查项名称", "检查结果", "问题描述", "位置", "修改建议"]
        for col, header in enumerate(headers, start=1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # 设置列宽
        ws_details.column_dimensions['A'].width = 8
        ws_details.column_dimensions['B'].width = 20
        ws_details.column_dimensions['C'].width = 25
        ws_details.column_dimensions['D'].width = 12
        ws_details.column_dimensions['E'].width = 40
        ws_details.column_dimensions['F'].width = 20
        ws_details.column_dimensions['G'].width = 30
        
        # 添加数据
        row_num = 2
        for result in doc_result.results:
            if result.issues:
                for issue in result.issues:
                    ws_details.cell(row=row_num, column=1, value=row_num-1).border = border
                    ws_details.cell(row=row_num, column=2, value=result.rule_name[:20] if len(result.rule_name) > 20 else result.rule_name).border = border
                    
                    # 获取规则类别
                    rule = doc_result.results[0].rule_id if doc_result.results else ""
                    ws_details.cell(row=row_num, column=3, value="技术文档规范").border = border
                    
                    status_cell = ws_details.cell(row=row_num, column=4, value="✗ 未通过")
                    status_cell.font = Font(color="FF0000")
                    status_cell.border = border
                    
                    ws_details.cell(row=row_num, column=5, value=issue.description).border = border
                    ws_details.cell(row=row_num, column=6, value=issue.position).border = border
                    ws_details.cell(row=row_num, column=7, value=issue.suggestion or "").border = border
                    
                    row_num += 1
            else:
                # 通过的项
                ws_details.cell(row=row_num, column=1, value=row_num-1).border = border
                ws_details.cell(row=row_num, column=2, value=result.rule_name[:20] if len(result.rule_name) > 20 else result.rule_name).border = border
                ws_details.cell(row=row_num, column=3, value="技术文档规范").border = border

                status_cell = ws_details.cell(row=row_num, column=4, value="✓ 通过")
                status_cell.font = Font(color="00B050")
                status_cell.border = border

                # 通过项不显示说明
                ws_details.cell(row=row_num, column=5, value="").border = border
                ws_details.cell(row=row_num, column=6, value="").border = border
                ws_details.cell(row=row_num, column=7, value="").border = border

                row_num += 1
        
        # 创建按类别统计工作表
        ws_category = wb.create_sheet("按类别统计")
        
        # 统计各类别的通过情况
        category_stats = {}
        for result in doc_result.results:
            # 从规则ID推断类别（简化处理）
            if 'cover' in result.rule_id or 'signature' in result.rule_id:
                category = "文档封面和签署页"
            elif 'page' in result.rule_id:
                category = "页码"
            elif 'toc' in result.rule_id:
                category = "目次"
            elif 'reference' in result.rule_id:
                category = "引用文件"
            elif 'text' in result.rule_id:
                category = "正文"
            elif 'figure' in result.rule_id or 'table' in result.rule_id:
                category = "图表"
            elif 'formula' in result.rule_id:
                category = "公式"
            elif 'unit' in result.rule_id or 'math' in result.rule_id or 'deviation' in result.rule_id:
                category = "量、单位及其符号"
            elif 'appendix' in result.rule_id:
                category = "附录"
            else:
                category = "其他"
            
            if category not in category_stats:
                category_stats[category] = {'total': 0, 'passed': 0, 'failed': 0}
            
            category_stats[category]['total'] += 1
            if result.passed:
                category_stats[category]['passed'] += 1
            else:
                category_stats[category]['failed'] += 1
        
        # 写入类别统计
        ws_category['A1'] = "类别统计"
        ws_category['A1'].font = Font(bold=True, size=14)
        
        headers = ["类别", "总检查项", "通过项", "失败项", "通过率"]
        for col, header in enumerate(headers, start=1):
            cell = ws_category.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row_num = 4
        for category, stats in category_stats.items():
            ws_category.cell(row=row_num, column=1, value=category).border = border
            ws_category.cell(row=row_num, column=2, value=stats['total']).border = border
            ws_category.cell(row=row_num, column=3, value=stats['passed']).border = border
            ws_category.cell(row=row_num, column=4, value=stats['failed']).border = border
            
            pass_rate = (stats['passed'] / stats['total'] * 100) if stats['total'] > 0 else 0
            ws_category.cell(row=row_num, column=5, value=f"{pass_rate:.1f}%").border = border
            
            row_num += 1
        
        # 保存文件
        wb.save(output_path)
    
    @staticmethod
    def export_to_html(doc_result: DocumentCheckResult, output_path: str):
        """
        导出为HTML格式
        
        Args:
            doc_result: 文档检查结果
            output_path: 输出文件路径
        """
        html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Word文档检查报告</title>
    <style>
        body {{
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        .header h1 {{
            margin: 0 0 10px 0;
        }}
        .header .info {{
            opacity: 0.9;
            font-size: 14px;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .stat-card {{
            background: white;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            text-align: center;
        }}
        .stat-card .number {{
            font-size: 36px;
            font-weight: bold;
            margin: 10px 0;
        }}
        .stat-card.passed .number {{
            color: #4CAF50;
        }}
        .stat-card.failed .number {{
            color: #f44336;
        }}
        .stat-card.warning .number {{
            color: #FF9800;
        }}
        .stat-card .label {{
            color: #666;
            font-size: 14px;
        }}
        .section {{
            background: white;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .section h2 {{
            margin-top: 0;
            color: #333;
            border-bottom: 2px solid #667eea;
            padding-bottom: 10px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }}
        th {{
            background-color: #667eea;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            padding: 12px;
            border-bottom: 1px solid #ddd;
        }}
        tr:hover {{
            background-color: #f5f5f5;
        }}
        .status-passed {{
            color: #4CAF50;
            font-weight: bold;
        }}
        .status-failed {{
            color: #f44336;
            font-weight: bold;
        }}
        .issue {{
            background-color: #fff3cd;
            padding: 15px;
            margin: 10px 0;
            border-left: 4px solid #ffc107;
            border-radius: 4px;
        }}
        .issue h4 {{
            margin: 0 0 10px 0;
            color: #856404;
        }}
        .footer {{
            text-align: center;
            color: #999;
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📄 Word文档检查报告</h1>
        <div class="info">
            <p><strong>文档路径：</strong>{doc_result.document_path}</p>
            <p><strong>检查时间：</strong>{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
        </div>
    </div>
    
    <div class="stats-grid">
        <div class="stat-card">
            <div class="label">总检查项</div>
            <div class="number">{doc_result.total_rules}</div>
        </div>
        <div class="stat-card passed">
            <div class="label">通过项</div>
            <div class="number">{doc_result.passed_rules}</div>
        </div>
        <div class="stat-card failed">
            <div class="label">失败项</div>
            <div class="number">{doc_result.failed_rules}</div>
        </div>
        <div class="stat-card warning">
            <div class="label">警告项</div>
            <div class="number">{doc_result.warning_rules}</div>
        </div>
    </div>
    
    <div class="section">
        <h2>📊 检查结果概览</h2>
        <table>
            <thead>
                <tr>
                    <th>序号</th>
                    <th>检查项名称</th>
                    <th>检查结果</th>
                    <th>问题数</th>
                    <th>说明</th>
                </tr>
            </thead>
            <tbody>
"""
        
        # 添加检查结果
        for i, result in enumerate(doc_result.results, start=1):
            status_class = "status-passed" if result.passed else "status-failed"
            status_text = "✓ 通过" if result.passed else "✗ 未通过"
            
            html_content += f"""
                <tr>
                    <td>{i}</td>
                    <td>{result.rule_name}</td>
                    <td class="{status_class}">{status_text}</td>
                    <td>{result.get_issue_count()}</td>
                    <td>{"" if result.passed else "存在问题"}</td>
                </tr>
"""
        
        html_content += """
            </tbody>
        </table>
    </div>
    
    <div class="section">
        <h2>🔍 问题详情</h2>
"""
        
        # 添加问题详情
        if doc_result.get_failed_results():
            for result in doc_result.get_failed_results():
                for issue in result.issues:
                    html_content += f"""
        <div class="issue">
            <h4>📌 {result.rule_name}</h4>
            <p><strong>位置：</strong>{issue.position}</p>
            <p><strong>问题：</strong>{issue.description}</p>
            {f'<p><strong>建议：</strong>{issue.suggestion}</p>' if issue.suggestion else ''}
        </div>
"""
        else:
            html_content += """
        <p style="color: #4CAF50; font-size: 18px; text-align: center; padding: 40px;">
            ✅ 所有问题均已通过检查！
        </p>
"""
        
        html_content += """
    </div>
    
    <div class="footer">
        <p>Word文档检查器 v2.0 | 生成时间：""" + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + """</p>
    </div>
</body>
</html>
"""
        
        # 保存文件
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    @staticmethod
    def export_to_json(doc_result: DocumentCheckResult, output_path: str):
        """
        导出为JSON格式
        
        Args:
            doc_result: 文档检查结果
            output_path: 输出文件路径
        """
        result_dict = {
            'document_path': doc_result.document_path,
            'check_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'total_rules': doc_result.total_rules,
            'passed_rules': doc_result.passed_rules,
            'failed_rules': doc_result.failed_rules,
            'warning_rules': doc_result.warning_rules,
            'pass_rate': doc_result.get_pass_rate(),
            'total_issues': doc_result.get_total_issues(),
            'results': []
        }
        
        for result in doc_result.results:
            result_item = {
                'rule_id': result.rule_id,
                'rule_name': result.rule_name,
                'passed': result.passed,
                'status': result.status.value if hasattr(result.status, 'value') else str(result.status),
                'message': "" if result.passed else result.message,  # 通过项不显示说明
                'issues': []
            }

            for issue in result.issues:
                result_item['issues'].append({
                    'position': issue.position,
                    'description': issue.description,
                    'suggestion': issue.suggestion
                })

            result_dict['results'].append(result_item)
        
        # 保存文件
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result_dict, f, ensure_ascii=False, indent=2)
